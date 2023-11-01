########################################################################################################################
# Filename: rc_database_tool.py
# Python Version 3.10
# Author: Sean Pack
########################################################################################################################
import argparse
# File Tools
import os
from pathlib import Path
import json
import openpyxl
from openpyxl.styles import PatternFill
# Access Database Tools
from win32com.client import Dispatch
import pypyodbc
# Python Data Manipulation
import pandas as pd

CONFIG_NAME = "config.json"

PATH = str(Path("").parent.absolute().parent.absolute())
DATA = PATH + "\\data\\"
DB = PATH + "\\databases\\"
EXPORTS = PATH + "\\exports\\"


def cli_run(args):
    if args.export_all_students:
        db = LocalDB(args.export_all_students[0])
        db.export_all_students()


class LocalDB:
    def __init__(self, filename):
        self.db_name = ""
        if filename.split('.')[1] == "accdb":
            self.db_name = filename
            if not os.path.exists(DATA + CONFIG_NAME):
                config = {}
            else:
                with open(DATA + CONFIG_NAME, "r") as fp:
                    config = json.load(fp)

            if self.db_name not in config:
                if not os.path.exists(DB + self.db_name):
                    print("%s file does not exist." % self.db_name)
                else:
                    print("%s exists." % self.db_name)
                config[self.db_name] = {"export_params": {},}
                with open(DATA + CONFIG_NAME, "w") as fp:
                    json.dump(config, fp, indent=4)
        else:
            print("%s file type not supported." % filename)

    def table_exists(self, table):
        dbname = DB + self.db_name
        constr = "DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={0};".format(dbname)

        dbconn = pypyodbc.connect(constr)

        cur = dbconn.cursor()
        try:
            cur.execute("SELECT * from [%s]" % table)
            return True
        except:
            return False

    def export_all_students(self):
        if os.path.exists(DB + self.db_name):
            if self.table_exists("All Students"):
                self.color_all_students()
            else:
                print("All Students table not in %s." % self.db_name)
        else:
            print("%s does not exist." % self.db_name)

    def add_export_param(self, column, unique_column, u_col_val, operator="<", val_1=None, val_2=None, color="00FFFF00"):
        valid_ops = ["<", "<=", ">", ">=", "< value <", "<= value <=", "> value >", ">= value >="]
        if operator in valid_ops:
            with open(DATA + CONFIG_NAME, "r") as fp:
                config = json.load(fp)

            df = self.get_db_table("All Students", self.db_name)
            unique_entries = df[unique_column].unique().tolist()

            if u_col_val in unique_entries:
                if column not in config[self.db_name]["export_params"]:
                    config[self.db_name]["export_params"][column] = []
                tmp_dict = {"u_col": unique_column, "u_val": u_col_val, "operator": operator}
                if val_1 is not None:
                    tmp_dict["val_1"] = val_1
                if val_2 is not None:
                    tmp_dict["val_2"] = val_2
                tmp_dict["color"] = color

                config[self.db_name]["export_params"][column].append(tmp_dict)
            else:
                print("%s is not a value in column %s" % (u_col_val, unique_column))
            with open(DATA + CONFIG_NAME, "w") as fp:
                json.dump(config, fp, indent=4)
        else:
            print("%s is not a valid operator." % operator)

    def color_all_students(self):
        # TODO REMOVE
        # export_params is assumed to be made this is a test
        # self.add_export_param("SOL LAST YEAR", "SUBJECT", "MATH_6", val_1=400, color="00FF0000")
        # self.add_export_param("SOL LAST YEAR", "SUBJECT", "MATH_6",
        #                       operator="< value <", val_1=400, val_2=425)
        # ------------------------------------

        exp_path = EXPORTS + "\\%s\\" % self.db_name
        if not os.path.exists(exp_path):
            os.mkdir(exp_path)
        exp_path += "all_students.xlsx"
        with open(DATA + CONFIG_NAME, "r") as fp:
            config = json.load(fp)
        df = self.get_db_table("All Students", self.db_name)
        df.to_excel(exp_path, index=False)
        wb = openpyxl.load_workbook(exp_path)
        first_sheet = wb.sheetnames[0]
        ws = wb[first_sheet]
        # Worksheet Column Names Translation Layer
        col_names = {}
        current = 0
        for col in ws.iter_cols(1, ws.max_column):
            col_names[col[0].value] = current
            current += 1
        e_params = config[self.db_name]["export_params"]

        for col in e_params:
            for command in e_params[col]:
                self.color_switch(ws, col_names, col, command)

        self.autofit_ws(ws)

        wb.save(exp_path)

    def all_students_join(self, table_name):
        # Only joins with all students
        pass

    @staticmethod
    def autofit_ws(ws):
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def color_switch(ws, col_names, col, command):
        if command["operator"] == "<":
            for row_cell in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row_cell[col_names[command["u_col"]]].value == command["u_val"]:
                    if int(row_cell[col_names[col]].value) < command["val_1"]:
                        row_cell[col_names[col]].fill = PatternFill(patternType="solid", fgColor=command["color"])
        if command["operator"] == "<=":
            for row_cell in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row_cell[col_names[command["u_col"]]].value == command["u_val"]:
                    if int(row_cell[col_names[col]].value) <= command["val_1"]:
                        row_cell[col_names[col]].fill = PatternFill(patternType="solid", fgColor=command["color"])
        if command["operator"] == ">":
            for row_cell in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row_cell[col_names[command["u_col"]]].value == command["u_val"]:
                    if int(row_cell[col_names[col]].value) > command["val_1"]:
                        row_cell[col_names[col]].fill = PatternFill(patternType="solid", fgColor=command["color"])
        if command["operator"] == ">=":
            for row_cell in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row_cell[col_names[command["u_col"]]].value == command["u_val"]:
                    if int(row_cell[col_names[col]].value) >= command["val_1"]:
                        row_cell[col_names[col]].fill = PatternFill(patternType="solid", fgColor=command["color"])
        if command["operator"] == "< value <":
            for row_cell in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row_cell[col_names[command["u_col"]]].value == command["u_val"]:
                    if command["val_1"] < int(row_cell[col_names[col]].value) < command["val_2"]:
                        row_cell[col_names[col]].fill = PatternFill(patternType="solid", fgColor=command["color"])
        if command["operator"] == "<= value <=":
            for row_cell in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row_cell[col_names[command["u_col"]]].value == command["u_val"]:
                    if command["val_1"] <= int(row_cell[col_names[col]].value) <= command["val_2"]:
                        row_cell[col_names[col]].fill = PatternFill(patternType="solid", fgColor=command["color"])
        if command["operator"] == "> value >":
            for row_cell in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row_cell[col_names[command["u_col"]]].value == command["u_val"]:
                    if command["val_1"] > int(row_cell[col_names[col]].value) > command["val_2"]:
                        row_cell[col_names[col]].fill = PatternFill(patternType="solid", fgColor=command["color"])
        if command["operator"] == ">= value >=":
            for row_cell in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row_cell[col_names[command["u_col"]]].value == command["u_val"]:
                    if command["val_1"] >= int(row_cell[col_names[col]].value) >= command["val_2"]:
                        row_cell[col_names[col]].fill = PatternFill(patternType="solid", fgColor=command["color"])

    @staticmethod
    def df_to_columns(df):
        column_string = ""
        cols = df.columns.tolist()
        for idx, d in enumerate(df.dtypes):
            column_string += "[%s] " % cols[idx]
            if d == "int64":
                column_string += "Int"
            elif d == "object":
                column_string += "Text"
            else:
                # Don't know what could happen here
                column_string += "Text"
            if not idx == len(cols) - 1:
                column_string += ", "
        return column_string

    @staticmethod
    def create_access_file(name):
        try:
            dbname = DB + name
            acc_app = Dispatch("Access.Application")
            db_engine = acc_app.DBEngine
            workspace = db_engine.Workspaces(0)

            db_lang_general = ';LANGID=0x0409;CP=1252;COUNTRY=0'
            new_db = workspace.CreateDatabase(dbname, db_lang_general, 64)

        except Exception as e:
            print(e)

        # finally:
        #     accApp.DoCmd.CloseDatabase
        #     accApp.Quit
        #     new_db = None
        #     workspace = None
        #     db_engine = None
        #     acc_app = None

    @staticmethod
    def execute_db(command, name):
        dbname = DB + name
        constr = "DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={0};".format(dbname)

        dbconn = pypyodbc.connect(constr)

        cur = dbconn.cursor()
        cur.execute(command)
        dbconn.commit()

    @staticmethod
    def get_db_table(table, name):
        dbname = DB + name
        constr = "DRIVER={{Microsoft Access Driver (*.mdb, *.accdb)}};DBQ={0};".format(dbname)

        dbconn = pypyodbc.connect(constr)

        cur = dbconn.cursor()
        cur.execute("SELECT * from [%s]" % table)
        columns = [column[0] for column in cur.description]
        for idx, c in enumerate(columns):
            columns[idx] = c.upper()

        df = pd.DataFrame(cur.fetchall(), columns=columns)
        return df


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Database Tool CLI")
    parser.add_argument("-e", "--export_all_students",
                        action="store",
                        metavar="filename",
                        nargs=1,
                        help="Name of file to import. Name of target database.")
    cli_run(parser.parse_args())
